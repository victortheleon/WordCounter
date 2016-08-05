from openpyxl import load_workbook
import os
from ignored_words import STOP_WORDS


FILE_PATH = os.path.abspath("../data/input.xlsx")


def split_date():
    """
    Simple method to split the date in workbook and save it to different location
    :return:
    """
    wb = load_workbook(FILE_PATH)
    ws = wb.worksheets[0]
    ws['G1'] = 'date'
    ws['H1'] = 'time'
    for i in range(1, len(ws.columns[4])):
        cell_value = ws.columns[4][i]
        ws['G' + str(1+i)] = cell_value.value.split('  ')[0]
        ws['H' + str(1+i)] = cell_value.value.split('  ')[1]
    wb.save(os.path.abspath("../data/output.xlsx"))


def plot_histogram():
    current_path = os.path.abspath("../data/output.xlsx")
    wb = load_workbook(current_path)
    wb.create_sheet("results", 0)
    result_ws = wb.active
    data_ws = wb.worksheets[1]
    histogram = dict()
    for cell in data_ws.columns[2][1:]:
        cell_words = cell.value.lower().split()
        for word in cell_words:
            if word not in STOP_WORDS:
                if word not in histogram:
                    histogram[word] = 1
                else:
                    histogram[word] += 1
    result_ws['A1'] = "word"
    result_ws['B1'] = "count"
    counter = 2
    for item in histogram:
        result_ws['A' + str(counter)] = item
        result_ws['B' + str(counter)] = histogram[item]
        counter += 1
    result_ws.auto_filter.ref = "A1:" + result_ws.columns[1][-1].coordinate
    wb.save(current_path)

split_date()
plot_histogram()
